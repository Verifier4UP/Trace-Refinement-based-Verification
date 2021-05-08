#!/usr/bin/env python
#-*- coding:utf-8 -*-
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


# line chart
def plot(read_file):

    # plt.figure(figsize=(10, 6))
    x = np.linspace(1, 4, 200)
    y = x
    plt.plot(x, y, c='b', ls='-')


    wb_read = load_workbook(read_file)
    wb_read1 = wb_read.active
    x = []
    y = []
    i = 2
    while True:
        program = wb_read1.cell(row=i, column=1).value
        UA_result = wb_read1.cell(row=i, column=3).value
        UA_timecost = str(wb_read1.cell(row=i, column=4).value)
        main_result = wb_read1.cell(row=i, column=6).value
        main_timecost = str(wb_read1.cell(row=i, column=7).value)
        if program == None:
            break
        else:
            if "correct" in UA_result and "correct" in main_result:
                x.append(float(UA_timecost))
                y.append(float(main_timecost))
        i += 1

    N = len(x)
    count = 0
    for i in range(N):
        if x[i] > y[i]:
            count += 1
    print("UAutomizer: ", x)
    print("Ours:       ", y)
    print("Total: ", N, "programs")
    print("Ours is better: ", count, "programs")

    ratio = []
    for i in range(N):
        ratio.append('%.2f' % (x[i]/y[i]))
    ratio.sort()
    for i in range(N):
        print(i+1, ratio[i])



    colors = np.random.rand(N)  # Randomly generate 50 color values between 0 and 1
    # area = np.pi * (10 * np.random.rand(N)) ** 2  # Point radius range: 0~10
    area = 50
    plt.scatter(x, y, s=area, c=colors, alpha=0.5,
                marker=(9, 3, 30))
    # s is the area size of the point
    # alpha is the transparency of the color of the point
    # marker is the shape of the designated point marker (numsides, style, angle)

    plt.legend()
    plt.xticks(fontsize=15)
    plt.yticks(fontsize=15)
    plt.xlabel(u"UAutomizer", fontsize=20)
    plt.ylabel("Ours", fontsize=20)
    plt.title("ratio", fontsize=20)
    plt.show()
    # plt.savefig("Ratio.pdf")
    plt.close()


if __name__ == "__main__":
    read_file = "collect-svcomp-1-600.xlsx"
    plot(read_file)