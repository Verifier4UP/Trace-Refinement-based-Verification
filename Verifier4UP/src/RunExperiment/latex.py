import sys
from openpyxl import load_workbook

def latex(read_file, speed_ratio_all_flag, speed_co_nco_flag):

    latex_text = ""

    wb_read = load_workbook(read_file)
    wb_read1 = wb_read.active

    i = 2
    UA_timeout_num = 0
    main_timeout_num = 0
    UA_better_num = 0
    main_better_num = 0
    UA_better_num_union = 0
    main_better_num_union = 0
    success_num = 0
    speed_ratio = []
    co_speed_ratio = []
    nco_speed_ratio = []
    while True:
        # ----------------------------collect information------------------------------------------- #
        program = wb_read1.cell(row=i, column=1).value
        LoC = wb_read1.cell(row=i, column=2).value
        UA_result = wb_read1.cell(row=i, column=3).value
        UA_timecost = str(wb_read1.cell(row=i, column=4).value)
        UA_iteration = wb_read1.cell(row=i, column=5).value
        main_result = wb_read1.cell(row=i, column=6).value
        main_timecost = str(wb_read1.cell(row=i, column=7).value)
        main_transformation_time = str(wb_read1.cell(row=i, column=8).value)
        main_seperation_time =str(wb_read1.cell(row=i, column=9).value)
        main_Verify_time = str(wb_read1.cell(row=i, column=10).value)
        main_CEGAR_time = str(wb_read1.cell(row=i, column=11).value)
        main_detail = wb_read1.cell(row=i, column=12).value
        if main_detail != None:
            coherent_num = main_detail[:main_detail.rfind("\\")]
            non_coherent_num = main_detail[main_detail.rfind("\\")+1: main_detail.find("(")]
        if program == None:
            break
        else:
            program = program.replace("_", "\_")
            i += 1
        if not speed_co_nco_flag: main_timecost = str('%.3f' % (float(main_timecost) - float(main_transformation_time)))


        # ---------------------------------check consistence----------------------------------------- #
        if (UA_result == "incorrect" and main_result == "correct") or (UA_result == "correct" and main_result == "incorrect"):
            program = '\cellcolor {red}$' + program + '$'


        # ---------------------------------check timeout/error--------------------------------------- #
        if not "correct" in UA_result:
            UA_result = '\cellcolor {yellow}$' + UA_result + '$'
            UA_timeout_num += 1
            UA_result = UA_result.replace("timeout", "TO")
        if not "correct" in main_result:
            main_result = '\cellcolor {yellow}$' + main_result + '$'
            main_timeout_num += 1
            main_result = main_result.replace("timeout", "TO")


        # ---------------------------------calculate speed ratio------------------------------------- #
        speed_ratio_all = speed_ratio_all_flag
        speed_co_nco = speed_co_nco_flag
        if "yellow" in main_result and "correct" in UA_result:
            if speed_ratio_all: speed_ratio.append(float(UA_timecost) / float(600))
        elif "yellow" in UA_result and "correct" in main_result:
            if speed_ratio_all: speed_ratio.append(float(600) / float(main_timecost))
        elif not("yellow" in main_result) and not("yellow" in UA_result):
            # speed_ratio.append(float(UA_timecost) / float(main_timecost))
            if float(UA_timecost) > float(main_timecost):
                speed_ratio.append(float(UA_timecost) / float(main_timecost))
            else:
                speed_ratio.append(-float(main_timecost) / float(UA_timecost))
            if speed_co_nco:
                if float(main_Verify_time) == 0.0:
                    # nco_speed_ratio.append(float(UA_timecost) / float(main_timecost))
                    if float(UA_timecost) > float(main_timecost):
                        nco_speed_ratio.append(float(UA_timecost) / float(main_timecost))
                    else:
                        nco_speed_ratio.append(-float(main_timecost) / float(UA_timecost))
                elif float(main_CEGAR_time) == 0.0:
                    # co_speed_ratio.append(float(UA_timecost) / float(main_timecost))
                    if float(UA_timecost) > float(main_timecost):
                        co_speed_ratio.append(float(UA_timecost) / float(main_timecost))
                    else:
                        co_speed_ratio.append(-float(main_timecost) / float(UA_timecost))
        else:
            if speed_ratio_all: speed_ratio.append(1)


        # ---------------------------------compare performance---------------------------------------- #
        if "yellow" in main_result and "correct" in UA_result:
            UA_timecost = '\cellcolor{gray!60}$' + UA_timecost + '$'
            UA_better_num += 1
        elif "yellow" in UA_result and "correct" in main_result:
            main_timecost = '\cellcolor{gray!60}$' + main_timecost + '$'
            main_better_num += 1
        elif not("yellow" in main_result) and not("yellow" in UA_result):
            if float(UA_timecost) > float(main_timecost):
                main_timecost = '\cellcolor{gray!60}$' + main_timecost + '$'
                main_better_num += 1
                main_better_num_union += 1
                success_num += 1
            elif float(UA_timecost) < float(main_timecost):
                UA_timecost = '\cellcolor{gray!60}$' + UA_timecost + '$'
                UA_better_num += 1
                UA_better_num_union += 1
                success_num += 1

        if "yellow" in main_result:
            main_result = "TO"
            main_timecost = "TO"
            if float(main_Verify_time) == 0.0:
                main_Verify_time = "TO"
            if float(main_CEGAR_time) == 0.0:
                main_CEGAR_time = "TO"
        if "yellow" in UA_result:
            UA_result = "TO"
            UA_timecost = "TO"

        latex_text = latex_text + \
                     program + \
                     " & " + LoC + \
                     " & " + UA_result + \
                     " & " + UA_timecost + \
                     " & " + main_result + \
                     " & " + main_timecost + \
                     " & " + main_seperation_time + \
                     " & " + main_Verify_time + "(" + coherent_num + ")"  \
                     " & " + main_CEGAR_time + "(" + non_coherent_num + ")" + "\\\\\n"

    # ---------------------------------output result---------------------------------------- #
    # print("\n\n\n\hline\n" + latex_text + "\hline\n\n")
    print("\n\n")
    print("There are total " + str(i-2) + " programs.")
    print("UAutomizer: timeout-" + str(UA_timeout_num) + "  better-" + str(UA_better_num))
    print("Ours: timeout-" + str(main_timeout_num) + " better-" + str(main_better_num))
    print("for both succeed program(" + str(success_num) + "), UAutomizer is better(" + str(UA_better_num_union) + "), Ours is better(" + str(main_better_num_union) + ")")
    print("and the speed_ratio is ", str(sum(speed_ratio)/len(speed_ratio)) + "(" + str(len(speed_ratio)) + ")")
    if speed_co_nco: print("and co_speed_ratio is ", str(sum(co_speed_ratio)/len(co_speed_ratio)) + "(" + str(len(co_speed_ratio)) + ")")
    if speed_co_nco: print("and nco_speed_ratio is ", str(sum(nco_speed_ratio) / len(nco_speed_ratio)) + "(" + str(len(nco_speed_ratio)) + ")")



    # ---------------------------------for fig.dat---------------------------------------- #
    with open("./fig.dat", 'w') as f:
        speed_ratio.sort()
        for i in range(len(speed_ratio)):
            f.write(str(i+1) + " " + str(speed_ratio[i]) + "\n")


if __name__ == "__main__":

    # read_file = "collect-benchmark-1-600.xlsx"
    # speed_ratio_all_flag = "False"
    # speed_co_nco_flag = "False"

    # read_file = "collect-svcomp-1-600.xlsx"
    # speed_ratio_all_flag = "False"
    # speed_co_nco_flag = "True"


    read_file = sys.argv[1]
    speed_ratio_all_flag = sys.argv[2]
    speed_co_nco_flag = sys.argv[3]
    if speed_ratio_all_flag == "False":
        speed_ratio_all_flag = False
    else:
        speed_ratio_all_flag = True
    if speed_co_nco_flag == "False":
        speed_co_nco_flag = False
    else:
        speed_co_nco_flag = True
    latex(read_file, speed_ratio_all_flag, speed_co_nco_flag)
