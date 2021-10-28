import os
import sys
from openpyxl import Workbook, load_workbook

if __name__ == "__main__":

    file = sys.argv[1]
    # file = "collect.xlsx"

    if os.path.exists(file):
        os.system("rm " + file)
    write_file = file
    wb = Workbook()
    wb.save(write_file)
    wb_write = load_workbook(write_file)
    wb_write1 = wb_write.active
    wb_write1.cell(row=1, column=1, value="program")
    wb_write1.cell(row=1, column=2, value="LoC")
    wb_write1.cell(row=1, column=3, value="UA-result")
    wb_write1.cell(row=1, column=4, value="UA-timecost")
    wb_write1.cell(row=1, column=5, value="UA-iteration")
    wb_write1.cell(row=1, column=6, value="main-result")
    wb_write1.cell(row=1, column=7, value="main-timecost")
    wb_write1.cell(row=1, column=8, value="main-trans")
    wb_write1.cell(row=1, column=9, value="main-sep")
    wb_write1.cell(row=1, column=10, value="main-V")
    wb_write1.cell(row=1, column=11, value="main-T")
    wb_write1.cell(row=1, column=12, value="(#V)/(#T)")


    current_line = 2
    total_num = 0
    for root, dirs, files in os.walk("./"):
        dirs.sort()
        for dir in dirs:
            if dir[-1].isdigit():
                total_num += 1
    for i in range(total_num):
        # Traverse all folders
        if os.path.exists("./RunExperiment" + str(i) + "/result_ours.txt"):
            with open("./RunExperiment" + str(i) + "/result_ours.txt", 'r') as f:
                for line in f:
                    if "bpl" in line:
                        line = line.split()
                        wb_write1.cell(row=current_line, column=1, value=line[0][:line[0].rfind(".bpl")])
                        wb_write1.cell(row=current_line, column=2, value=line[1])
                        wb_write1.cell(row=current_line, column=6, value=line[2])
                        wb_write1.cell(row=current_line, column=7, value=float(line[3]))
                        wb_write1.cell(row=current_line, column=8, value=float(line[4]))
                        wb_write1.cell(row=current_line, column=9, value=float(line[5]))
                        wb_write1.cell(row=current_line, column=10, value=float(line[6]))
                        wb_write1.cell(row=current_line, column=11, value=float(line[7]))
                        wb_write1.cell(row=current_line, column=12, value=line[8])
        if os.path.exists("./RunExperiment" + str(i) + "/result_ultimate.txt"):
            with open("./RunExperiment" + str(i) + "/result_ultimate.txt") as f:
                for line in f:
                    if "bpl" in line:
                        line = line.split()
                        wb_write1.cell(row=current_line, column=1, value=line[0][:line[0].rfind(".bpl")])
                        wb_write1.cell(row=current_line, column=2, value=line[1])
                        wb_write1.cell(row=current_line, column=3, value=line[2])
                        wb_write1.cell(row=current_line, column=4, value=float(line[3]))
                        wb_write1.cell(row=current_line, column=5, value=line[4])
        current_line += 1
    wb_write.save(write_file)